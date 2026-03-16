from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.orm import Session
from database import get_db
from models import TrainingPlan, PlannedSession, SessionResult
from pydantic import BaseModel
from typing import Optional
import openpyxl
import io
from datetime import datetime

router = APIRouter(prefix="/training", tags=["training"])


# ── Schemas ────────────────────────────────────────────────────────────────────

class SessionResultCreate(BaseModel):
    status: str                        # completed | partial | skipped
    duration_real: Optional[float] = None
    distance_real: Optional[float] = None
    avg_pulse: Optional[int] = None
    feeling: Optional[str] = None     # excellent | good | regular | hard
    notes: Optional[str] = None


# ── Helpers ────────────────────────────────────────────────────────────────────

def parse_float(val):
    try:
        return float(val) if val not in (None, "", "LIBRE") else None
    except:
        return None

def parse_int(val):
    try:
        return int(float(val)) if val not in (None, "") else None
    except:
        return None

def is_rest_day(exercise_type):
    if not exercise_type:
        return True
    return str(exercise_type).strip().upper() == "LIBRE"


# ── Upload & Parse Excel ───────────────────────────────────────────────────────

@router.post("/upload/{user_id}")
async def upload_training_plan(user_id: str, file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Solo se aceptan archivos .xlsx")

    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents))

    # Leer Config
    if "Config" not in wb.sheetnames:
        raise HTTPException(status_code=400, detail="El archivo no tiene hoja 'Config'. Usa la plantilla oficial.")

    cfg = wb["Config"]
    def cfg_val(row): 
        return cfg.cell(row=row, column=3).value

    plan_name    = str(cfg_val(4) or "Sin nombre")
    objective    = str(cfg_val(10) or "")
    competition  = str(cfg_val(9) or "")
    observations = str(cfg_val(12) or "")
    fecha_inicio = str(cfg_val(7) or "")

    # Determinar mes y año desde fecha_inicio
    month, year = datetime.now().month, datetime.now().year
    try:
        parts = fecha_inicio.replace("-", "/").split("/")
        if len(parts) == 3:
            month, year = int(parts[1]), int(parts[2])
    except:
        pass

    # Verificar si ya existe un plan para ese mes/año
    existing = db.query(TrainingPlan).filter(
        TrainingPlan.user_id == user_id,
        TrainingPlan.month == month,
        TrainingPlan.year == year
    ).first()
    if existing:
        db.delete(existing)
        db.commit()

    # Crear plan
    plan = TrainingPlan(
        user_id=user_id,
        name=plan_name,
        month=month,
        year=year,
        competition=competition,
        objective=objective,
        observations=observations
    )
    db.add(plan)
    db.flush()

    # Leer hoja Rutina
    if "Rutina" not in wb.sheetnames:
        raise HTTPException(status_code=400, detail="El archivo no tiene hoja 'Rutina'. Usa la plantilla oficial.")

    rut = wb["Rutina"]
    sessions_created = 0

    for row in rut.iter_rows(min_row=3, values_only=True):
        # Columnas: semana, fecha, dia, tipo, descripcion, duracion, distancia, pulso_min, pulso_max, notas
        if all(v is None for v in row):
            continue

        week          = parse_int(row[0])
        date          = str(row[1]) if row[1] else None
        day           = str(row[2]) if row[2] else None
        exercise_type = str(row[3]) if row[3] else None
        description   = str(row[4]) if row[4] else None
        duration_min  = parse_float(row[5])
        distance_km   = parse_float(row[6])
        pulse_min     = parse_int(row[7])
        pulse_max     = parse_int(row[8])
        notes         = str(row[9]) if row[9] else None

        session = PlannedSession(
            plan_id=plan.id,
            week=week,
            date=date,
            day=day,
            exercise_type=exercise_type,
            description=description,
            duration_min=duration_min,
            distance_km=distance_km,
            pulse_min=pulse_min,
            pulse_max=pulse_max,
            notes=notes,
            is_rest=is_rest_day(exercise_type)
        )
        db.add(session)
        sessions_created += 1

    db.commit()

    return {
        "message": f"Plan cargado exitosamente. {sessions_created} sesiones importadas.",
        "plan_id": plan.id,
        "month": month,
        "year": year
    }


# ── Get plans (historial por usuario) ─────────────────────────────────────────

@router.get("/plans/{user_id}")
def get_plans(user_id: str, db: Session = Depends(get_db)):
    plans = db.query(TrainingPlan).filter(
        TrainingPlan.user_id == user_id
    ).order_by(TrainingPlan.year.desc(), TrainingPlan.month.desc()).all()

    return [
        {
            "id": p.id,
            "name": p.name,
            "month": p.month,
            "year": p.year,
            "competition": p.competition,
            "objective": p.objective,
            "uploaded_at": p.uploaded_at
        }
        for p in plans
    ]


# ── Get sessions de un plan ────────────────────────────────────────────────────

@router.get("/plan/{plan_id}/sessions")
def get_sessions(plan_id: int, db: Session = Depends(get_db)):
    sessions = db.query(PlannedSession).filter(
        PlannedSession.plan_id == plan_id
    ).order_by(PlannedSession.date).all()

    return [session_to_dict(s) for s in sessions]


# ── Get sesiones del mes actual ────────────────────────────────────────────────

@router.get("/current/{user_id}")
def get_current_plan(user_id: str, db: Session = Depends(get_db)):
    now = datetime.now()
    plan = db.query(TrainingPlan).filter(
        TrainingPlan.user_id == user_id,
        TrainingPlan.month == now.month,
        TrainingPlan.year == now.year
    ).first()

    if not plan:
        return {"plan": None, "sessions": []}

    sessions = db.query(PlannedSession).filter(
        PlannedSession.plan_id == plan.id
    ).order_by(PlannedSession.date).all()

    return {
        "plan": {
            "id": plan.id,
            "name": plan.name,
            "month": plan.month,
            "year": plan.year,
            "competition": plan.competition,
            "objective": plan.objective,
            "observations": plan.observations
        },
        "sessions": [session_to_dict(s) for s in sessions]
    }


# ── Registrar resultado de una sesión ─────────────────────────────────────────

@router.post("/session/{session_id}/result/{user_id}")
def save_result(session_id: int, user_id: str, payload: SessionResultCreate, db: Session = Depends(get_db)):
    session = db.query(PlannedSession).filter(PlannedSession.id == session_id).first()
    if not session:
        raise HTTPException(status_code=404, detail="Sesión no encontrada")

    existing = db.query(SessionResult).filter(
        SessionResult.session_id == session_id,
        SessionResult.user_id == user_id
    ).first()

    if existing:
        for k, v in payload.dict().items():
            setattr(existing, k, v)
        existing.recorded_at = datetime.now()
    else:
        result = SessionResult(
            session_id=session_id,
            user_id=user_id,
            **payload.dict()
        )
        db.add(result)

    db.commit()
    return {"message": "Resultado guardado"}

# ── Get todos los resultados de sesiones del usuario ──────────────────────────

@router.get("/sessions/results/{user_id}")
def get_all_session_results(user_id: str, db: Session = Depends(get_db)):
    results = (
        db.query(SessionResult, PlannedSession)
        .join(PlannedSession, SessionResult.session_id == PlannedSession.id)
        .filter(SessionResult.user_id == user_id)
        .order_by(SessionResult.recorded_at.desc())
        .all()
    )
    return [
        {
            "id": r.SessionResult.id,
            "status": r.SessionResult.status,
            "duration_real": r.SessionResult.duration_real,
            "distance_real": r.SessionResult.distance_real,
            "avg_pulse": r.SessionResult.avg_pulse,
            "feeling": r.SessionResult.feeling,
            "recorded_at": r.SessionResult.recorded_at,
            "exercise_type": r.PlannedSession.exercise_type,
            "description": r.PlannedSession.description,
            "date": r.PlannedSession.date,
        }
        for r in results
    ]


# ── Helper ─────────────────────────────────────────────────────────────────────

def session_to_dict(s: PlannedSession):
    result = None
    if s.result:
        result = {
            "id": s.result.id,
            "status": s.result.status,
            "duration_real": s.result.duration_real,
            "distance_real": s.result.distance_real,
            "avg_pulse": s.result.avg_pulse,
            "feeling": s.result.feeling,
            "notes": s.result.notes,
            "recorded_at": s.result.recorded_at
        }
    return {
        "id": s.id,
        "week": s.week,
        "date": s.date,
        "day": s.day,
        "exercise_type": s.exercise_type,
        "description": s.description,
        "duration_min": s.duration_min,
        "distance_km": s.distance_km,
        "pulse_min": s.pulse_min,
        "pulse_max": s.pulse_max,
        "notes": s.notes,
        "is_rest": s.is_rest,
        "result": result
    }